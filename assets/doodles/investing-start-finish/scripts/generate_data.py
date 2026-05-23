#!/usr/bin/env python3
"""Generate annual S&P 500 + Gold return data from Damodaran historical dataset."""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from urllib.request import urlopen

from openpyxl import load_workbook

DATA_URL = "https://www.stern.nyu.edu/~adamodar/pc/datasets/histretSP.xlsx"
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "data" / "annual_returns.json"


def read_inflation_map(workbook):
    ws = workbook["Inflation Rate"]
    inflation = {}
    for row in range(12, ws.max_row + 1):
        year = ws.cell(row, 1).value
        rate = ws.cell(row, 3).value
        if isinstance(year, int) and isinstance(rate, (int, float)):
            inflation[year] = float(rate)
    return inflation


def read_asset_returns(workbook):
    ws = workbook["Returns by year"]
    out = []
    for row in range(21, ws.max_row + 1):
        year = ws.cell(row, 1).value
        sp500_nominal = ws.cell(row, 2).value
        gold_nominal = ws.cell(row, 8).value
        if not isinstance(year, int):
            continue
        if not isinstance(sp500_nominal, (int, float)):
            continue
        gold_value = float(gold_nominal) if isinstance(gold_nominal, (int, float)) else None
        out.append((year, float(sp500_nominal), gold_value))
    return out


def main():
    content = urlopen(DATA_URL, timeout=30).read()
    workbook = load_workbook(BytesIO(content), data_only=True)

    inflation = read_inflation_map(workbook)
    asset_returns = read_asset_returns(workbook)

    series = []
    for year, sp500_nominal, gold_nominal in asset_returns:
        if year not in inflation:
            continue
        infl = inflation[year]
        sp500_real = (1.0 + sp500_nominal) / (1.0 + infl) - 1.0
        gold_real = None
        if isinstance(gold_nominal, float):
            gold_real = (1.0 + gold_nominal) / (1.0 + infl) - 1.0
        series.append(
            {
                "year": year,
                "sp500NominalReturn": sp500_nominal,
                "sp500RealReturn": sp500_real,
                "goldNominalReturn": gold_nominal,
                "goldRealReturn": gold_real,
                "inflation": infl,
            }
        )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": DATA_URL,
        "description": "Annual S&P 500 and gold returns with inflation-adjusted real return variants.",
        "startYear": series[0]["year"],
        "endYear": series[-1]["year"],
        "series": series,
    }
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH} ({len(series)} years)")


if __name__ == "__main__":
    main()
